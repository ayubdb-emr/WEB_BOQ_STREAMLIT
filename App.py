import streamlit as st
import openpyxl
from openpyxl import load_workbook
import gdown
import os
import zipfile
import re
import xml.etree.ElementTree as ET
import io
from shapely.geometry import LineString, Point
import networkx as nx
from pyproj import Geod

# Konfigurasi Halaman & Sembunyikan Menu Bawaan
st.set_page_config(
    page_title="EMR Tools Dashboard",
    page_icon="🛠️",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': None,
        'Report a bug': None,
        'About': None
    }
)

# Sembunyikan Header dan Footer Streamlit

geod = Geod(ellps="WGS84")

# ================= CONFIG & HELPER FUNCTIONS =================

MAX_DISTANCE = 0.005
RR_HH_THRESHOLD = 0.00003
HC_SLING_THRESHOLD = 0.00003

FDT_COLUMN_MAP = {
    1: "C", 2: "D", 3: "E", 4: "F", 5: "G",
    6: "H", 7: "I", 8: "J", 9: "K", 10: "L",
}

STANDARD_ROW_MAP = {
    ("A", "FO 24C/2T"): 2, ("B", "FO 24C/2T"): 3,
    ("C", "FO 24C/2T"): 4, ("D", "FO 24C/2T"): 5,
    ("A", "FO 36C/3T"): 6, ("B", "FO 36C/3T"): 7,
    ("C", "FO 36C/3T"): 8, ("D", "FO 36C/3T"): 9,
    ("A", "FO 48C/4T"): 10, ("B", "FO 48C/4T"): 11,
    ("C", "FO 48C/4T"): 12, ("D", "FO 48C/4T"): 13,
}

STRAND_WIRE_ROW = 15

FDT_TYPE_ROW = {
    "48": 38,
    "72": 39,
    "96": 40,
    "144": 41,
}

TOTAL_FAT_ROW = {
    "A": 45,
    "B": 46,
    "C": 47,
    "D": 48,
}

POLE_ROW = {
    "7M5": 74,
    "7M4": 75,
    "7M3": 76,
    "7M25": 77,
    "9M5": 78,
    "9M4": 79,
    "LN": 80,
    "MTI": 81,
    "MR": 82,
}

NRO_SHEET_NAME = "BoQ NRO Cluster"
NRO_CLUSTER_CELL = "O3"
NRO_HP_CELL = "O5"

CABLE_COLOR_MAP = {
    "FO 24C/2T": "#00ff00",
    "FO 36C/3T": "#ff00ff",
    "FO 48C/4T": "#aa00ff",
    "FO 72C/6T": "#0000ff",
}

def get_fdt_capacity_by_fat(total_fat):
    if total_fat <= 20:
        return "FDT 48C"
    elif total_fat <= 30:
        return "FDT 72C"
    elif total_fat <= 40:
        return "FDT 96C"
    elif total_fat <= 60:
        return "FDT 144C"
    elif total_fat <= 120:
        return "FDT 288C"
    else:
        return f"FDT >288C ({total_fat})"

def get_cable_type_and_color_by_fat(fat_count):
    if fat_count <= 10:
        cable_type = "FO 24C/2T"
    elif fat_count <= 15:
        cable_type = "FO 36C/3T"
    elif fat_count <= 20:
        cable_type = "FO 48C/4T"
    else:
        cable_type = "FO 72C/6T"
    return cable_type, CABLE_COLOR_MAP.get(cable_type)

def set_text(pm, tag, value):
    el = pm.find(tag)
    if el is None:
        el = ET.SubElement(pm, tag)
    el.text = str(value)

def geodesic_length(coords):
    total = 0.0
    for i in range(len(coords) - 1):
        lon1, lat1 = coords[i]
        lon2, lat2 = coords[i + 1]
        _, _, dist = geod.inv(lon1, lat1, lon2, lat2)
        total += dist
    return total

def extract_coords_from_linestring(pm):
    ls = pm.find(".//LineString")
    if ls is None:
        return []
    coord_el = ls.find("coordinates")
    if coord_el is None or not coord_el.text:
        return []
    coords = []
    for c in coord_el.text.strip().split():
        lon, lat, *_ = map(float, c.split(","))
        coords.append((lon, lat))
    return coords

def extract_point_from_placemark(pm):
    pt = pm.find(".//Point")
    if pt is None:
        return None
    coord_el = pt.find(".//coordinates")
    if coord_el is None or not coord_el.text:
        return None
    lon, lat, *_ = map(float, coord_el.text.strip().split(","))
    return Point(lon, lat)

def norm_text(text):
    return (text or "").upper().replace('"', "").replace("'", "")

def has_any(text, variants):
    t = norm_text(text)
    return any(v in t for v in variants)

def is_text_7_25(text):
    text = str(text).upper()
    variants = ["7-2.5", "7 - 2.5", "7- 2.5", "7 -2.5", "7/2.5", "7 / 2.5", "7M2.5", "7M 2.5", "7 M 2.5"]
    return any(v in text for v in variants)

def is_text_7_3(text):
    return has_any(text, ["7-3", "7 / 3", "7M3", "7M 3", "7 M 3"])

def is_text_7_4(text):
    return has_any(text, ["7-4", "7 / 4", "7M4", "7M 4", "7 M 4"])

def is_text_9_5(text):
    return has_any(text, ["9-5", "9 / 5", "9M5", "9M 5", "9 M 5"])

def is_text_9_4(text):
    return has_any(text, ["9-4", "9 / 4", "9M4", "9M 4", "9 M 4"])

def extract_line_code(line_name):
    m = re.search(r"LINE\s+([A-Z0-9]+)", line_name.upper())
    if m:
        return m.group(1)
    return line_name.strip().upper()

def extract_fdt_index_from_line_name(line_name, total_fdt=1):
    m = re.search(r"FDT\s*(\d+)", line_name.upper())
    if m:
        return int(m.group(1))
    if total_fdt == 1:
        return 1
    return None

def get_line_folders(root):
    result = []
    for folder in root.findall(".//Folder"):
        name = folder.findtext("name", "").strip()
        if name.upper().startswith("LINE "):
            result.append((name, folder))
    return result

def auto_detect_fdt_codes(root):
    codes = []
    for folder in root.findall(".//Folder"):
        fname = folder.findtext("name", "").strip().upper()
        if fname != "FDT":
            continue
        for pm in folder.findall("Placemark"):
            if pm.find(".//Point") is None:
                continue
            name = pm.findtext("name", "").strip()
            if name:
                codes.append(name)
    unique = []
    for c in codes:
        if c not in unique:
            unique.append(c)
    return unique

def get_fdt_placemarks(root):
    fdt_list = []
    for folder in root.findall(".//Folder"):
        fname = folder.findtext("name", "").strip().upper()
        if fname != "FDT":
            continue
        for pm in folder.findall("Placemark"):
            if pm.find(".//Point") is not None:
                fdt_list.append(pm)
    return fdt_list

def get_fdt_code_by_index(fdt_codes, fdt_index):
    if fdt_index is None:
        return None
    idx = fdt_index - 1
    if 0 <= idx < len(fdt_codes):
        return fdt_codes[idx]
    return None

def find_nearest_fat_name(target_point, fat_items):
    nearest_name = None
    nearest_dist = float("inf")
    for fat_name, fat_point in fat_items:
        dist = target_point.distance(fat_point)
        if dist < nearest_dist:
            nearest_dist = dist
            nearest_name = fat_name
    return nearest_name

def line_passes_near_any_point(line_coords, point_items, threshold=0.00003):
    line = LineString(line_coords)
    for _, pt in point_items:
        if line.distance(pt) <= threshold:
            return True
    return False

def point_near_any_line(point, line_items, threshold=0.00003):
    for line in line_items:
        if line.distance(point) <= threshold:
            return True
    return False

def build_graph_from_line_folder(line_folder):
    G = nx.Graph()
    start_node = None
    cable_line_placemarks = []
    for folder in line_folder.findall(".//Folder"):
        fname = folder.findtext("name", "").upper()
        if "DISTRIBUTION CABLE" in fname or "SLING WIRE" in fname:
            for pm in folder.findall("Placemark"):
                coords = extract_coords_from_linestring(pm)
                if len(coords) < 2:
                    continue
                if "DISTRIBUTION CABLE" in fname:
                    cable_line_placemarks.append(pm)
                for i in range(len(coords) - 1):
                    a = coords[i]
                    b = coords[i + 1]
                    seg = LineString([a, b])
                    G.add_edge(a, b, geometry=seg)
                if start_node is None:
                    start_node = coords[0]
    return G, start_node, cable_line_placemarks

def build_virtual_line(G, start_node, reverse_input="n"):
    ordered_nodes = list(nx.dfs_preorder_nodes(G, start_node))
    if reverse_input == "y":
        ordered_nodes.reverse()
    path_coords = []
    for i in range(len(ordered_nodes) - 1):
        a = ordered_nodes[i]
        b = ordered_nodes[i + 1]
        edge = G.get_edge_data(a, b)
        if not edge:
            continue
        coords = list(edge["geometry"].coords)
        if coords[0] != a:
            coords.reverse()
        if path_coords and path_coords[-1] == coords[0]:
            path_coords.extend(coords[1:])
        else:
            path_coords.extend(coords)
    return LineString(path_coords), path_coords

def collect_objects_from_line_folder(line_folder, virtual_line):
    all_objects = []
    for folder in line_folder.findall(".//Folder"):
        folder_name = folder.findtext("name", "")
        fname = folder_name.upper()
        for pm in folder.findall("Placemark"):
            pt = pm.find(".//Point")
            if pt is None:
                continue
            coord_el = pt.find(".//coordinates")
            if coord_el is None or not coord_el.text:
                continue
            lon, lat, *_ = map(float, coord_el.text.strip().split(","))
            point = Point(lon, lat)
            if virtual_line.distance(point) > MAX_DISTANCE:
                continue
            chain = virtual_line.project(point)
            if "NEW POLE" in fname:
                obj = "NEW"
            elif "EXISTING POLE EMR" in fname:
                obj = "EMR"
            elif "EXISTING POLE PARTNER" in fname:
                obj = "PARTNER"
            elif "FAT" in fname and "BOUNDARY" not in fname:
                obj = "FAT"
            elif "NEW HH 20X20X20" in fname:
                obj = "HHPIT"
            elif "NEW HH 40X40X60" in fname:
                obj = "HH"
            elif ("HANDHOLE" in fname or "HH" in fname) and "BOUNDARY" not in fname:
                obj = "HH"
            else:
                continue
            all_objects.append((chain, pm, obj, folder_name))
    all_objects.sort(key=lambda x: x[0])
    return all_objects

def get_partner_prefix(old_name):
    old = old_name.upper().strip()
    if any(k in old for k in ["TEL", "TA", "TLKM", "TELKOM"]):
        return "EXT.TLKM"
    elif "PLN" in old:
        return "EXT.PLN"
    elif "MTI" in old:
        return "EXT.MTI"
    elif any(k in old for k in ["FM", "FIRSTMEDIA", "LN", "LINKNET"]):
        return "EXT.LN"
    else:
        return "EXT.PARTNER"

def get_cluster_name_from_file(filename):
    base = re.sub(r"\.(kmz|kml)$", "", filename, flags=re.IGNORECASE)
    base = re.sub(r"\s*\(\d+\)$", "", base).strip()
    if " - " in base:
        return base.split(" - ", 1)[1].strip()
    return re.sub(r"^TGR\d+\s*-?\s*", "", base, flags=re.IGNORECASE).strip()

def count_hp_cover(root):
    count = 0
    counted_ids = set()
    for folder in root.findall(".//Folder"):
        folder_name = folder.findtext("name", "") or ""
        folder_upper = folder_name.upper()
        is_hp_folder = (
            "HP COVER" in folder_upper
            or "HP COVERAGE" in folder_upper
            or folder_upper.strip() in ["HP"]
        )
        for pm in folder.findall(".//Placemark"):
            pm_id = id(pm)
            name = pm.findtext("name", "") or ""
            desc = pm.findtext("description", "") or ""
            text_upper = f"{folder_name} {name} {desc}".upper()
            has_point = pm.find(".//Point") is not None
            if is_hp_folder and has_point:
                count += 1
                counted_ids.add(pm_id)
            elif pm_id not in counted_ids and has_point and (
                "HP COVER" in text_upper
                or "HP COVERAGE" in text_upper
                or "HOMEPASS" in text_upper
            ):
                count += 1
                counted_ids.add(pm_id)
    return count

def boq_put(ws, fdt_index, row, value):
    col = FDT_COLUMN_MAP.get(fdt_index)
    if not col:
        return
    if value in (None, "", 0):
        ws[f"{col}{row}"] = None
    else:
        ws[f"{col}{row}"] = value

def rgb_hex_to_kml_color(hex_color, alpha="ff"):
    hex_color = hex_color.strip().lower().replace("#", "")
    if not re.fullmatch(r"[0-9a-f]{6}", hex_color):
        return None
    rr = hex_color[0:2]
    gg = hex_color[2:4]
    bb = hex_color[4:6]
    return f"{alpha}{bb}{gg}{rr}"

def set_linestyle_color(pm, hex_color, width="3"):
    kml_color = rgb_hex_to_kml_color(hex_color)
    if not kml_color:
        return
    style_el = pm.find("Style")
    if style_el is None:
        style_el = ET.SubElement(pm, "Style")
    line_style = style_el.find("LineStyle")
    if line_style is None:
        line_style = ET.SubElement(style_el, "LineStyle")
    color_el = line_style.find("color")
    if color_el is None:
        color_el = ET.SubElement(line_style, "color")
    color_el.text = kml_color
    width_el = line_style.find("width")
    if width_el is None:
        width_el = ET.SubElement(style_el, "width")
    width_el.text = str(width)


# ================= NAVIGASI SIDEBAR =================

st.sidebar.title("Dashboard Menu")
menu = st.sidebar.radio("Pilih Tools:", ["BOQ Smart Downloader", "Multi FDT Processor"])


# ============================================================
# MENU 1: BOQ SMART DOWNLOADER
# ============================================================
if menu == "BOQ Smart Downloader":
    st.title("Aplikasi BOQ Only - Smart Downloader")
    st.write("Target Sheet: **BoM CLUSTER AE**")

    FILE_ID = "1bOWQbhFRuOINRlL6truxY3b-UKs6vWBm"
    boq_path = "BoQ_AE_Template.xlsx"

    @st.cache_resource
    def download_template():
        if os.path.exists(boq_path):
            os.remove(boq_path)
        url_drive = f"https://drive.google.com/uc?id={FILE_ID}"
        try:
            gdown.download(url_drive, boq_path, quiet=True)
        except Exception:
            pass
        if not os.path.exists(boq_path) or not zipfile.is_zipfile(boq_path):
            if os.path.exists(boq_path):
                os.remove(boq_path)
            url_export = f"https://docs.google.com/spreadsheets/d/{FILE_ID}/export?format=xlsx"
            try:
                gdown.download(url_export, boq_path, quiet=True)
            except Exception:
                pass

    download_template()

    if not os.path.exists(boq_path) or not zipfile.is_zipfile(boq_path):
        st.error("Gagal mendownload template BOQ dari Google Drive. Pastikan link Google Drive sudah public.")
    else:
        st.success("Template BOQ berhasil dimuat!")

    uploaded_kmz = st.file_uploader("Upload file As-plan drawing (.kmz / .kml):", type=['kmz', 'kml'], key="boq_upload")

    if uploaded_kmz is not None:
        if st.button("Proses Data BOQ"):
            with st.spinner("Sedang memproses data drawing..."):
                try:
                    kmz_path = uploaded_kmz.name
                    with open(kmz_path, "wb") as f:
                        f.write(uploaded_kmz.getbuffer())

                    if kmz_path.lower().endswith(".kmz"):
                        with zipfile.ZipFile(kmz_path, "r") as z:
                            kml_files = [f for f in z.namelist() if f.lower().endswith(".kml")]
                            if not kml_files:
                                st.error("File KMZ tidak berisi file KML yang valid.")
                                st.stop()
                            kml_file = kml_files[0]
                            raw = z.read(kml_file).decode("utf-8", "ignore")
                    else:
                        raw = open(kmz_path, encoding="utf-8").read()

                    raw = re.sub(r'\s+xmlns(:\w+)?="[^"]+"', "", raw)
                    raw = re.sub(r"<(/?)(\w+):", r"<\1", raw)
                    root = ET.fromstring(raw)

                    FDT_CODES = auto_detect_fdt_codes(root)
                    line_folders = get_line_folders(root)
                    cluster_name = get_cluster_name_from_file(kmz_path)
                    hp_cover_total = count_hp_cover(root)

                    st.write(f"**Cluster:** {cluster_name}")
                    st.write(f"**HP Cover Total:** {hp_cover_total}")

                    fat_count_by_fdt = {}
                    boq_standard_odn = {}
                    boq_strand_wire = {}
                    boq_total_fat_line = {}
                    boq_pole = {}

                    for line_name, line_folder in line_folders:
                        CODE_FAT = extract_line_code(line_name)
                        line_letter = CODE_FAT[0].upper()
                        fdt_index = extract_fdt_index_from_line_name(line_name, len(FDT_CODES))
                        if fdt_index is None:
                            fdt_index = 1

                        boq_strand_wire.setdefault(fdt_index, [])
                        boq_pole.setdefault(fdt_index, {
                            "7M5": 0, "7M4": 0, "7M3": 0, "7M25": 0,
                            "9M5": 0, "9M4": 0, "LN": 0, "MTI": 0, "MR": 0,
                        })

                        fat_count_by_fdt.setdefault(fdt_index, 0)
                        G, start_node, _ = build_graph_from_line_folder(line_folder)
                        if start_node is None or len(G.nodes) == 0:
                            continue

                        virtual_line, path_coords = build_virtual_line(G, start_node)
                        if len(path_coords) < 2:
                            continue

                        route_length_report = round(geodesic_length(path_coords))
                        all_objects = collect_objects_from_line_folder(line_folder, virtual_line)

                        fat_count = 0
                        sw_lengths = []

                        for chain, pm, obj, folder_name in all_objects:
                            if obj == "NEW":
                                pole_text = f"{pm.findtext('name','')} {pm.findtext('description','')} {folder_name}"
                                if is_text_7_25(pole_text):
                                    boq_pole[fdt_index]["7M25"] += 1
                                elif is_text_7_3(pole_text):
                                    boq_pole[fdt_index]["7M3"] += 1
                                elif is_text_7_4(pole_text):
                                    boq_pole[fdt_index]["7M4"] += 1
                                elif is_text_9_5(pole_text):
                                    boq_pole[fdt_index]["9M5"] += 1
                                elif is_text_9_4(pole_text):
                                    boq_pole[fdt_index]["9M4"] += 1
                                else:
                                    boq_pole[fdt_index]["7M5"] += 1
                            elif obj == "EMR":
                                boq_pole[fdt_index]["MR"] += 1
                            elif obj == "PARTNER":
                                old = pm.findtext("name", "").strip()
                                prefix = get_partner_prefix(old)
                                if prefix == "EXT.LN":
                                    boq_pole[fdt_index]["LN"] += 1
                                else:
                                    boq_pole[fdt_index]["MTI"] += 1
                            elif obj == "FAT":
                                fat_count += 1

                        fat_count_by_fdt[fdt_index] += fat_count

                        for folder in line_folder.findall(".//Folder"):
                            fname = folder.findtext("name", "").upper()
                            if "SLING WIRE" not in fname:
                                continue
                            for pm in folder.findall("Placemark"):
                                coords = extract_coords_from_linestring(pm)
                                if len(coords) < 2:
                                    continue
                                line_length_round = round(geodesic_length(coords))
                                sw_lengths.append(line_length_round)

                        boq_strand_wire[fdt_index].extend(sw_lengths)

                        if fat_count <= 10:
                            cable_type = "FO 24C/2T"
                        elif fat_count <= 15:
                            cable_type = "FO 36C/3T"
                        else:
                            cable_type = "FO 48C/4T"

                        boq_standard_odn[(fdt_index, line_letter, cable_type)] = route_length_report
                        boq_total_fat_line[(fdt_index, line_letter)] = fat_count

                    wb = load_workbook(boq_path)
                    if "BoM CLUSTER AE" not in wb.sheetnames:
                        st.error("Sheet 'BoM CLUSTER AE' tidak ditemukan pada file Excel!")
                    else:
                        ws = wb["BoM CLUSTER AE"]
                        rows_to_clear = (
                            list(range(2, 14))
                            + [STRAND_WIRE_ROW]
                            + list(range(38, 42))
                            + list(range(45, 49))
                            + list(range(74, 83))
                        )

                        for fdt_idx, col in FDT_COLUMN_MAP.items():
                            for row in rows_to_clear:
                                ws[f"{col}{row}"] = None

                        for (fdt_index, line_letter, cable_type), value in boq_standard_odn.items():
                            row = STANDARD_ROW_MAP.get((line_letter, cable_type))
                            if row:
                                boq_put(ws, fdt_index, row, value)

                        for fdt_index, lengths in boq_strand_wire.items():
                            col = FDT_COLUMN_MAP.get(fdt_index)
                            if not col:
                                continue
                            if lengths:
                                ws[f"{col}{STRAND_WIRE_ROW}"] = "=" + "+".join(str(x) for x in lengths)
                            else:
                                ws[f"{col}{STRAND_WIRE_ROW}"] = None

                        for fdt_index, total_fat in fat_count_by_fdt.items():
                            col = FDT_COLUMN_MAP.get(fdt_index)
                            if not col:
                                continue
                            if total_fat <= 20:
                                ws[f"{col}{FDT_TYPE_ROW['48']}"] = 1
                            elif total_fat <= 30:
                                ws[f"{col}{FDT_TYPE_ROW['72']}"] = 1
                            elif total_fat <= 40:
                                ws[f"{col}{FDT_TYPE_ROW['96']}"] = 1
                            else:
                                ws[f"{col}{FDT_TYPE_ROW['144']}"] = 1

                        for (fdt_index, line_letter), value in boq_total_fat_line.items():
                            row = TOTAL_FAT_ROW.get(line_letter)
                            if row:
                                boq_put(ws, fdt_index, row, value)

                        for fdt_index, data in boq_pole.items():
                            for key, row in POLE_ROW.items():
                                boq_put(ws, fdt_index, row, data.get(key, 0))

                        if NRO_SHEET_NAME in wb.sheetnames:
                            ws_nro = wb[NRO_SHEET_NAME]
                            ws_nro[NRO_CLUSTER_CELL] = cluster_name
                            ws_nro[NRO_HP_CELL] = hp_cover_total if hp_cover_total > 0 else None

                        output = io.BytesIO()
                        wb.save(output)

                        st.success("Proses BOQ Berhasil!")
                        st.download_button(
                            label="Download Hasil BoQ_AE_HASIL.xlsx",
                            data=output.getvalue(),
                            file_name="BoQ_AE_HASIL.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                except Exception as e:
                    st.error(f"Terjadi kesalahan saat memproses data: {e}")


# ============================================================
# MENU 2: MULTI FDT PROCESSOR
# ============================================================
elif menu == "Multi FDT Processor":
    st.title("Multi FDT Processor - KMZ/KML Utility")
    st.write("Alat otomatisasi pemrosesan jalur network, penomoran tiang, FAT, Handhole, Sling Wire, dan Kabel.")

    fdt_codes_raw = st.text_input("MASUKKAN CODE FDT (pisahkan dengan koma):", "FDT.01")
    FDT_CODES = [x.strip() for x in fdt_codes_raw.split(",") if x.strip()]

    if not FDT_CODES:
        st.error("Code FDT tidak boleh kosong. Harap masukkan minimal 1 kode (Contoh: FDT.01).")
        st.stop()

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        NEW_START = st.number_input("Start NEW POLE", value=1, min_value=1)
    with col2:
        EMR_START = st.number_input("Start EXT POLE", value=1, min_value=1)
    with col3:
        PARTNER_START = st.number_input("Start EXT PARTNER", value=1, min_value=1)
    with col4:
        FAT_START = st.number_input("Start FAT", value=1, min_value=1)

    reverse_input = st.selectbox("Balik arah jalur?", ["n", "y"])

    uploaded_file = st.file_uploader("Upload file KMZ atau KML:", type=["kmz", "kml"], key="fdt_upload")

    if uploaded_file is not None:
        if st.button("Proses File KMZ/KML"):
            with st.spinner("Sedang memproses struktur data jaringan..."):
                try:
                    file_bytes = uploaded_file.read()
                    
                    if uploaded_file.name.lower().endswith(".kmz"):
                        with zipfile.ZipFile(io.BytesIO(file_bytes), "r") as z:
                            kml_files = [f for f in z.namelist() if f.lower().endswith(".kml")]
                            if not kml_files:
                                st.error("File KMZ tidak berisi file KML.")
                                st.stop()
                            kml_file = kml_files[0]
                            raw = z.read(kml_file).decode("utf-8", "ignore")
                    else:
                        raw = file_bytes.decode("utf-8", "ignore")

                    raw = re.sub(r'\s+xmlns(:\w+)?="[^"]+"', "", raw)
                    raw = re.sub(r"<(/?)(\w+):", r"<\1", raw)
                    root = ET.fromstring(raw)

                    line_folders = get_line_folders(root)
                    if not line_folders:
                        st.error("Tidak ditemukan folder LINE pada file.")
                    else:
                        fdt_placemarks = get_fdt_placemarks(root)
                        if len(FDT_CODES) < len(fdt_placemarks):
                            st.warning(f"Jumlah code FDT kurang. Ditemukan {len(fdt_placemarks)} FDT di file, tapi input hanya {len(FDT_CODES)} code.")

                        new_counter_by_fdt = {}
                        emr_counter_by_fdt = {}
                        partner_counter_by_fdt = {}
                        hh_pit_counter_by_fdt = {}
                        hh_counter_by_fdt = {}
                        fat_count_by_fdt = {}

                        for line_name, line_folder in line_folders:
                            CODE_FAT = extract_line_code(line_name)
                            fat_no = FAT_START

                            fdt_index = extract_fdt_index_from_line_name(line_name, len(FDT_CODES))
                            CODE_FDT_CURRENT = get_fdt_code_by_index(FDT_CODES, fdt_index)

                            if fdt_index is None:
                                fdt_index = 1
                            if CODE_FDT_CURRENT is None:
                                CODE_FDT_CURRENT = FDT_CODES[0]

                            new_counter_by_fdt.setdefault(fdt_index, NEW_START)
                            emr_counter_by_fdt.setdefault(fdt_index, EMR_START)
                            partner_counter_by_fdt.setdefault(fdt_index, PARTNER_START)
                            hh_pit_counter_by_fdt.setdefault(fdt_index, 1)
                            hh_counter_by_fdt.setdefault(fdt_index, 1)
                            fat_count_by_fdt.setdefault(fdt_index, 0)

                            G, start_node, cable_line_placemarks = build_graph_from_line_folder(line_folder)
                            virtual_line, path_coords = build_virtual_line(G, start_node, reverse_input)
                            route_length = geodesic_length(path_coords)
                            all_objects = collect_objects_from_line_folder(line_folder, virtual_line)

                            sling_lines_for_pole = []
                            for folder in line_folder.findall(".//Folder"):
                                fname = folder.findtext("name", "").upper()
                                if "SLING WIRE" not in fname:
                                    continue
                                for pm in folder.findall("Placemark"):
                                    coords = extract_coords_from_linestring(pm)
                                    if len(coords) < 2:
                                        continue
                                    sling_lines_for_pole.append(LineString(coords))

                            fat_count = 0
                            fat_objects_for_sling, hh_objects_for_sling, all_partner_objects = [], [], []

                            for chain, pm, obj, folder_name in all_objects:
                                if obj == "NEW":
                                    base_name = f"MR.{CODE_FDT_CURRENT.split('.')[0]}.P{new_counter_by_fdt[fdt_index]:03d}"
                                    pt = extract_point_from_placemark(pm)
                                    old_name = pm.findtext("name", "").upper()
                                    old_desc = pm.findtext("description", "").upper()
                                    folder_upper = folder_name.upper()

                                    is_new_pole_7_25 = is_text_7_25(old_name) or is_text_7_25(old_desc) or is_text_7_25(folder_upper)

                                    if is_new_pole_7_25 and pt is not None and point_near_any_line(pt, sling_lines_for_pole, threshold=HC_SLING_THRESHOLD):
                                        name = f"{base_name}"
                                    else:
                                        name = base_name

                                    new_counter_by_fdt[fdt_index] += 1
                                    set_text(pm, "name", name)
                                    desc_el = pm.find("description")
                                    if desc_el is not None:
                                        pm.remove(desc_el)

                                elif obj == "EMR":
                                    name = f"EXT-MR.{CODE_FDT_CURRENT.split('.')[0]}.P{emr_counter_by_fdt[fdt_index]:03d}"
                                    emr_counter_by_fdt[fdt_index] += 1
                                    set_text(pm, "name", name)
                                    desc_el = pm.find("description")
                                    if desc_el is not None:
                                        pm.remove(desc_el)

                                elif obj == "FAT":
                                    name = f"{CODE_FDT_CURRENT}.{CODE_FAT}{fat_no:02d}"
                                    set_text(pm, "name", name)
                                    set_text(pm, "description", f"FAT.{CODE_FAT}{fat_no:02d}")
                                    fat_count += 1
                                    fat_point = extract_point_from_placemark(pm)
                                    if fat_point is not None:
                                        fat_objects_for_sling.append((name, fat_point))
                                    fat_no += 1

                                elif obj == "HHPIT":
                                    name = f"MR.{CODE_FDT_CURRENT.split('.')[0]}.HHPIT{hh_pit_counter_by_fdt[fdt_index]:03d}"
                                    hh_pit_counter_by_fdt[fdt_index] += 1
                                    set_text(pm, "name", name)
                                    set_text(pm, "description", "NEW HH 20X20X20")
                                    hh_point = extract_point_from_placemark(pm)
                                    if hh_point is not None:
                                        hh_objects_for_sling.append((name, hh_point))

                                elif obj == "HH":
                                    name = f"MR.{CODE_FDT_CURRENT.split('.')[0]}.HH{hh_counter_by_fdt[fdt_index]:03d}"
                                    hh_counter_by_fdt[fdt_index] += 1
                                    set_text(pm, "name", name)
                                    set_text(pm, "description", "NEW HH 40X40X60")
                                    hh_point = extract_point_from_placemark(pm)
                                    if hh_point is not None:
                                        hh_objects_for_sling.append((name, hh_point))

                                elif obj == "PARTNER":
                                    all_partner_objects.append((chain, pm))

                            fat_count_by_fdt[fdt_index] += fat_count

                            for _, pm in sorted(all_partner_objects, key=lambda x: x[0]):
                                old = pm.findtext("name", "").strip()
                                prefix = get_partner_prefix(old)
                                new_name = f"{prefix}.{CODE_FDT_CURRENT.split('.')[0]}.P{partner_counter_by_fdt[fdt_index]:03d}"
                                partner_counter_by_fdt[fdt_index] += 1
                                set_text(pm, "name", new_name)
                                desc_el = pm.find("description")
                                if desc_el is not None:
                                    pm.remove(desc_el)

                            cable_type, cable_color = get_cable_type_and_color_by_fat(fat_count)
                            slack_fat = fat_count
                            slack_unit = 1 + slack_fat
                            slack_length = slack_unit * 20
                            tolerance = (route_length + slack_length) * 0.05
                            total_cable = route_length + slack_length + tolerance

                            route_length_report = round(route_length)
                            slack_length_report = round(slack_length)
                            tolerance_report = round(tolerance)
                            total_cable_report = round(total_cable)

                            cable_title = f"{CODE_FDT_CURRENT} - Cable Line {CODE_FAT} ({cable_type}) - AE - {total_cable_report}m"
                            cable_description = (
                                f"Nilai Route      : {route_length_report} m\n"
                                f"Total Slack      : {slack_unit} Unit (1 slack FDT & {slack_fat} slack FAT) @20m\n"
                                f"Toleransi 5%     : {tolerance_report} m\n\n"
                                f"Total Length Cable : {route_length_report} + {slack_length_report} + {tolerance_report} = {total_cable_report} m"
                            )

                            for pm in cable_line_placemarks:
                                set_text(pm, "name", cable_title)
                                set_text(pm, "description", cable_description)
                                if cable_color:
                                    set_linestyle_color(pm, cable_color, width="3")

                            for folder in line_folder.findall(".//Folder"):
                                fname = folder.findtext("name", "").upper()
                                if "SLING WIRE" not in fname:
                                    continue
                                for pm in folder.findall("Placemark"):
                                    coords = extract_coords_from_linestring(pm)
                                    if len(coords) < 2:
                                        continue
                                    line_length_round = round(geodesic_length(coords))

                                    start_point = Point(coords[0])
                                    end_point = Point(coords[-1])
                                    nearest_fat_start = find_nearest_fat_name(start_point, fat_objects_for_sling)
                                    nearest_fat_end = find_nearest_fat_name(end_point, fat_objects_for_sling)

                                    if nearest_fat_start is None and nearest_fat_end is None:
                                        continue

                                    dist_start, dist_end = float("inf"), float("inf")
                                    for fat_name, fat_point in fat_objects_for_sling:
                                        if fat_name == nearest_fat_start:
                                            dist_start = start_point.distance(fat_point)
                                        if fat_name == nearest_fat_end:
                                            dist_end = end_point.distance(fat_point)

                                    nearest_fat_name = nearest_fat_start if dist_start <= dist_end else nearest_fat_end
                                    if nearest_fat_name is None:
                                        continue

                                    fat_suffix = nearest_fat_name.split(".")[-1]
                                    is_rodding_rope = line_passes_near_any_point(coords, hh_objects_for_sling, threshold=RR_HH_THRESHOLD)

                                    if is_rodding_rope:
                                        sling_name = f"RR-{fat_suffix} - {line_length_round} m"
                                        sling_desc = "Tambang pancing/Roding rope"
                                    else:
                                        sling_name = f"SW-{fat_suffix}"
                                        sling_desc = f"{line_length_round} m"

                                    set_text(pm, "name", sling_name)
                                    set_text(pm, "description", sling_desc)

                            for folder in line_folder.findall(".//Folder"):
                                fname = folder.findtext("name", "").upper()
                                if "SLACK HANGER" not in fname:
                                    continue
                                for pm in folder.findall("Placemark"):
                                    pt = extract_point_from_placemark(pm)
                                    if pt is None:
                                        continue
                                    nearest_fat_name = find_nearest_fat_name(pt, fat_objects_for_sling)
                                    if nearest_fat_name is None:
                                        continue
                                    fat_suffix = nearest_fat_name.split(".")[-1]
                                    slack_name = f"SLACK-{fat_suffix}"
                                    set_text(pm, "name", slack_name)
                                    set_text(pm, "description", "SLACK HANGER")

                        for i, pm in enumerate(fdt_placemarks, start=1):
                            fdt_code = FDT_CODES[i - 1] if (i - 1) < len(FDT_CODES) else FDT_CODES[0]
                            total_fat = fat_count_by_fdt.get(i, 0)
                            fdt_desc = get_fdt_capacity_by_fat(total_fat)
                            set_text(pm, "name", fdt_code)
                            set_text(pm, "description", fdt_desc)

                        out_kml = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                        output_kmz = io.BytesIO()
                        with zipfile.ZipFile(output_kmz, "w", zipfile.ZIP_DEFLATED) as z:
                            z.writestr("doc.kml", out_kml)

                        st.success("File KMZ/KML Berhasil Diproses!")
                        st.download_button(
                            label="Download Hasil FINAL_MULTI_FDT.kmz",
                            data=output_kmz.getvalue(),
                            file_name="FINAL_MULTI_FDT.kmz",
                            mime="application/vnd.google-earth.kmz"
                        )
                except Exception as e:
                    st.error(f"Terjadi kesalahan saat memproses file KML/KMZ: {e}")
